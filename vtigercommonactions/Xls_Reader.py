import openpyxl





workbook = openpyxl.load_workbook("C:/Users/dell/PycharmProjects/vTigerPytestFramework/Files/TestData.xlsx")
sheet = workbook["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column

dict = {}
for r in range(2, rows+1):
    dict1 = {}
    for c in range(2, cols+1):
        #print(col[row].value)
        #dt = sheet.cell(3,4).value
        #print(dt)
         key = sheet.cell(1,c).value
         val = sheet.cell(r,c).value
         dict1[key]=val
         #print("key ="+key+" value="+val)

    dict[sheet.cell(r,1).value]= dict1

#print(dict)

print(dict.get("TC04").get("LastName"))



# dict1 = {"userid":"admin1","password":"admin1"}
# dict2 = {"userid":"XYZ","password":"admin2"}
# dict3 = {"userid":"admin3","password":"admin3"}
#
# dict = {"TC01":dict1,"TC02":dict2,"TC03":dict3}
#
# print(dict.get("TC02").get("userid"))


