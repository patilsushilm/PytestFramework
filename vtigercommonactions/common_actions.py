import time
import openpyxl

from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert

class common_actions:

    ewait = 2

    def __init__(self, driver):
        self.driver = driver


    def readExcel(self,excelpath,sheetName):
        workbook = openpyxl.load_workbook(excelpath)
        sheet = workbook[sheetName]
        rows = sheet.max_row
        cols = sheet.max_column
        dict = {}
        for r in range(2, rows + 1):
            dict1 = {}
            for c in range(2, cols + 1):
                key = sheet.cell(1, c).value
                val = sheet.cell(r, c).value
                dict1[key] = val

            dict[sheet.cell(r, 1).value] = dict1
        return dict


    def enter_text(self,locator,text):
        WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator)).clear()
        WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator)).send_keys(text)

    def click_element(self, locator):
        WebDriverWait(self.driver, self.ewait).until(EC.element_to_be_clickable(locator)).click()
        time.sleep(3)

    def select_text_by_index(self, locator, index):
        elm = WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator))
        select = Select(elm)
        select.select_by_index(index)

    def select_text_by_text(self, locator, text):
        elm = WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator))
        select = Select(elm)
        select.select_by_visible_text(text)

    def element_exist(self,locator):
        return WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator)).is_displayed()

    def get_element_text(self,locator):
        return WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator)).text

    def get_textbox_text(self,locator,attribute):
        return WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator)).get_attribute(attribute)

    def get_alert_text(self):
        alert = Alert(self.driver)
        return alert.text

    def alert_accept(self):
        alert = Alert(self.driver)
        alert.accept()

    def alert_dismiss(self):
        alert = Alert(self.driver)
        alert.dismiss()


    def move_mouse(self,locator):
        elm = WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator))
        act = ActionChains(self.driver)
        act.move_to_element(elm).perform()

    def move_left_click(self,locator):
        elm = WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(locator))
        act = ActionChains(self.driver)
        act.click(elm).perform()

    def mouse_drag_drop(self,source_locator,dest_locator):
        source = WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(source_locator))
        dest = WebDriverWait(self.driver, self.ewait).until(EC.visibility_of_element_located(dest_locator))
        act = ActionChains(self.driver)
        act.drag_and_drop(source,dest).perform()








